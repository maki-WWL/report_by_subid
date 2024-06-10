import pandas as pd
import os
import numpy as np

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from kt_comparison.services.checkers import WWLTraffWorldChecker, TraffManagerChecker, ThirdKtChecker


def should_be_deleted(row):
    for item in row:
        if pd.notna(item) and item != 0:
            return False
        elif pd.notna(item) and item == 0:
            return True
    return True


def create_excel(dates_of_files: str, first_date_str: str, second_date_str: str, folder_path: str) -> None:
    base_dir = os.path.join(folder_path)
    excel_writer = pd.ExcelWriter(f'result_{folder_path}.xlsx', engine='openpyxl')

    for date_of_file in dates_of_files:
        file_name = f"{date_of_file.split('-')[0]}-{date_of_file.split('-')[1]}.csv"

        first_file_path = os.path.join(base_dir, first_date_str, file_name)
        second_file_path = os.path.join(base_dir, second_date_str, file_name)

        df1 = pd.read_csv(first_file_path, quotechar='"', sep=';')
        df2 = pd.read_csv(second_file_path, quotechar='"', sep=';')
        print(df1.columns)
        df1['SubId'] = df1['SubId'].astype(str)
        df2['SubId'] = df2['SubId'].astype(str)

        df1.drop_duplicates(subset=['SubId'], inplace=True)
        df2.drop_duplicates(subset=['SubId'], inplace=True)

        # df1.set_index('SubId', inplace=True)
        # df2.set_index('SubId', inplace=True)

        combined = pd.merge(df1, df2, on='SubId', how='outer', suffixes=('_file1', '_file2'))

        if 'Revenue_file1' in combined.columns:
            combined.rename(columns={'Revenue_file1': 'Доход_file1'}, inplace=True)
        if 'Revenue_file2' in combined.columns:
            combined.rename(columns={'Revenue_file2': 'Доход_file2'}, inplace=True)

        only_in_df1 = combined[combined['Доход_file2'].isna()]
        only_in_df2 = combined[combined['Доход_file1'].isna()]

        combined['Доход_file1'] = pd.to_numeric(combined['Доход_file1'], errors='coerce')
        combined['Доход_file2'] = pd.to_numeric(combined['Доход_file2'], errors='coerce')

        diff = combined[(~combined['Доход_file1'].isna() & ~combined['Доход_file2'].isna()) & (combined['Доход_file1'] != combined['Доход_file2'])]

        diff['Різниця'] = diff['Доход_file2'].fillna(0) - diff['Доход_file1'].fillna(0)

        first_column_name = combined.columns[0]

        empty_row_1 = pd.DataFrame(np.nan, index=[0], columns=combined.columns)
        empty_row_1[first_column_name] = 'Є у першому файлі, немає в другому'
            
        empty_row_2 = pd.DataFrame(np.nan, index=[0], columns=combined.columns)
        empty_row_2[first_column_name] = 'Є у другому файлі, немає в першому'

        result = pd.concat([diff, empty_row_1, only_in_df1, empty_row_2, only_in_df2])

        result.to_csv('result.csv', sep=';', encoding='utf-8', index=False)

        df = pd.read_csv('result.csv', sep=';')

        df = df[~df.apply(should_be_deleted, axis=1)]

        #delete 23:59

        subids = df['SubId'].dropna().unique().tolist()

        # if folder_path == 'kt_1':
        #     delete_subid = []
        #     print(len(subids))
        #     tm_checker = TraffManagerChecker()
        #     delete_subid = tm_checker.get_delete_list(subids)
        #     print(len(delete_subid))

        if folder_path == 'kt_2':
            wt_checker = WWLTraffWorldChecker()
            delete_subid = wt_checker.get_delete_list(subids)

        if folder_path == 'kt_3':
            third_kt_checker = ThirdKtChecker()  
            delete_subid = third_kt_checker.get_delete_list(subids)      

        df = df[~df['SubId'].isin(delete_subid)]

        sheet_name = date_of_file

        df.to_excel(excel_writer, sheet_name=sheet_name, index=False)

        os.remove('result.csv')

    excel_writer.close()



def resize_table() -> None:
    wb = load_workbook('result.xlsx')
    ws = wb.active

    # Автоматично налаштовуємо ширину стовпців
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Отримуємо літеру стовпця

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save('result.xlsx')


if __name__ == "__main__":
    create_excel(['5-2023', '6-2023'], '2023-11-21', '2023-11-22')
    # resize_table()